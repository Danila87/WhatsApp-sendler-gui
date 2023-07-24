import re

from Excel_work import excel_table
from colorama import init, Fore
from openpyxl.styles import PatternFill

init(autoreset=True)


def replace_string(string: str) -> str:

    """
    Функция форматирует строку (Конкретно здесь - номер). Удаляет лишние символы, добавляет отсутствующие куски номера.

    :param string: Строка для редактирования

    :return: Возвращает переформатированную строку
    """

    # Удаление лишних символов
    string = re.sub(r',', ';', string)
    string = re.sub(r'[^\d;]', '', string)

    list_number = string.split(';')

    for i, phone in enumerate(list_number):

        result = re.match(r'\b\d{6}\b', phone)
        if result:
            list_number[i] = f'73952{phone}'

        result = re.search(r'^8', phone)
        if result:
            list_number[i] = re.sub(r'^8', '7', phone)

        result = re.search(r'^3', phone)
        if result:
            list_number[i] = f'7{phone}'

        list_number[i] = re.sub(r'\b\d{1,10}\b', '', list_number[i])

    list_number = list(set(list_number))
    
    string = (";".join(list_number))

    string = re.sub(r'^;', '', string)

    return string


def update_number():

    """
    Функция запускает процесс форматирования номера. Проходит по строкам с номерами и форматирует их. За форматирование
    отвечает функция replace_string(). В конечном итоге сохраняет файл
    """

    max_row = excel_table.get_sheet_row()
    columns = excel_table.selected_column

    for row in excel_table.sheet_main.iter_rows(min_row=2, max_row=max_row):
        for i in columns:
            if row[i].value is None:
                row[i].fill = PatternFill(fill_type='solid', start_color='c4c43f')
                continue

            val = str(row[i].value)
            val = replace_string(val)
            row[i].value = val
            if row[i].value is None:
                row[i].fill = PatternFill(fill_type='solid', start_color='c4c43f')

    excel_table.save_excel()

    print(f"\n{Fore.GREEN}Номера отредактированы!\n")
